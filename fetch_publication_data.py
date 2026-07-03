import requests, re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

beamlines = {
    '2-ID':    {'fullname': 'Soft Inelastic X-ray Scattering',
                'tla':      'SIX',
                'program':  'Electronic Structure Techniques',
                'status':   'operating'},
    
    '3-ID':    {'fullname': 'Hard X-ray Nanoprobe',
                'tla':      'HXN',
                'program':  'Imaging and Microscopy',
                'status':   'operating'},
    
    '4-BM':    {'fullname': 'X-ray Fluorescence Microscopy',
                'tla':      'XFM',
                'program':  'BioImaging',
                'status':   'operating'},
    
    '4-ID':    {'fullname': 'Integrated In situ and Resonant Hard X-ray Studies',
                'tla':      'ISR',
                'program':  'Complex Scattering',
                'status':   'operating'},
    
    '5-ID':    {'fullname': 'Submicron Resolution X-ray Spectroscopy',
                'tla':      'SRX',
                'program':  'Imaging and Microscopy',
                'status':   'operating'},
    
    '6-BM':    {'fullname': 'Beamline for Materials Measurement',
                'tla':      'BMM',
                'program':  'Spectroscopy',
                'status':   'operating'},
    
    '6-ID':    {'fullname': 'High Resolution Powder Diffraction',
                'tla':      'HRD',
                'program':  'Hard X-ray Methods',
                'status':   'development'},
    
    '7-BM':    {'fullname': 'Quick x-ray Absorption and Scattering',
                'tla':      'QAS',
                'program':  'Spectroscopy',
                'status':   'operating'},
    
    '7-ID-1':  {'fullname': 'Spectroscopy Soft and Tender',
                'tla':      'SST1',
                'program':  'Spectroscopy',
                'status':   'operating'},
    
    '7-ID-2':  {'fullname': 'Spectroscopy Soft and Tender 2',
                'tla':      'SST2',
                'program':  'Spectroscopy',
                'status':   'operating'},
    
    '8-BM':    {'fullname': 'Tender Energy X-ray Absorption Spectroscopy',
                'tla':      'TES',
                'program':  'Spectroscopy',
                'status':   'operating'},
    
    '8-ID':    {'fullname': 'Inner-Shell Spectroscopy',
                'tla':      'ISS',
                'program':  'Spectroscopy',
                'status':   'operating'},
    
    '9-ID':    {'fullname': 'Coherent Diffraction Imaging',
                'tla':      'CDI',
                'program':  'Imaging and Microscopy',
                'status':   'development'},
    
    '10-ID':   {'fullname': 'Inelastic X-ray Scattering',
                'tla':      'IXS',
                'program':  'Complex Scattering',
                'status':   'operating'},
    
    '11-BM':   {'fullname': 'Complex Materials Scattering',
                'tla':      'CMS',
                'program':  'Complex Scattering',
                'status':   'operating'},
    
    '11-ID':   {'fullname': 'Coherent Hard X-ray Scattering',
                'tla':      'CHX',
                'program':  'Complex Scattering',
                'status':   'operating'},
    
    '12-ID':   {'fullname': 'Soft Matter Interfaces',
                'tla':      'SMI',
                'program':  'Complex Scattering',
                'status':   'operating'},
    
    '13-ID':   {'fullname': 'Advanced Materials Processing',
                'tla':      'AMP',
                'program':  'Complex Scattering',
                'status':   'development'},
    
    '16-BM':   {'fullname': 'Quantitive Cellular Tomography',
                'tla':      'QCT',
                'program':  'Studies Biology',
                'status':   'development'},
    
    '16-ID':   {'fullname': 'Life Science X-ray Scattering',
                'tla':      'LiX',
                'program':  'Structural Biology',
                'status':   'operating'},
    
    '17-BM':   {'fullname': 'X-ray Footprinting of Biological Materials',
                'tla':      'XFP',
                'program':  'Structural Biology',
                'status':   'operating'},
    
    '17-ID-1': {'fullname': 'Highly Automated Macromolecular Crystallography',
                'tla':      'AMX',
                'program':  'Structural Biology',
                'status':   'operating'},
    
    '17-ID-2': {'fullname': 'Frontier Microfocusing Macromolecular Crystallography',
                'tla':      'FMX',
                'program':  'Structural Biology',
                'status':   'operating'},
    
    '18-ID':   {'fullname': 'Full Field X-ray Imaging',
                'tla':      'FXI',
                'program':  'Imaging and Microscopy',
                'status':   'operating'},
    
    '19-ID':   {'fullname': 'Biological Microdiffraction Facility',
                'tla':       'NYX',
                'program':  'Structural Biology',
                'status':   'operating'},
    
    '21-ID':   {'fullname': 'Electron Spectro-Microscopy',
                'tla':      'ESM',
                'program':  'Electronic Structure Techniques',
                'status':   'operating'},
    
    '22-IR-1': {'fullname': 'Frontier Synchrotron Infrared Spectroscopy',
                'tla':      'FIS',
                'program':  'Electronic Structure Techniques',
                'status':   'operating'},
    
    '22-IR-2': {'fullname': 'Magnetospectroscopy, Ellipsometry and Time-Resolved Optical Spectroscopie',
                'tla':      'MET',
                'program':  'Electronic Structure Techniques',
                'status':   'operating'},
    
    '23-ID-1': {'fullname': 'Coherent Soft X-ray Scattering beamline',
                'tla':      'CSX',
                'program':  'Electronic Structure Techniques',
                'status':   'operating'},
    
    '23-ID-2': {'fullname': 'In situ and Operando Soft X-ray Spectroscopy',
                'tla':      'IOS',
                'program':  'Spectroscopy',
                'status':   'operating'},
    
    '24-IR':   {'fullname': 'IR Nanospctroscopy Microspectroscopy',
                'tla':      'INF',
                'program':  'Electronic Structure Techniques',
                'status':   'development'},
    
    '26-ID-1': {'fullname': 'Advanced Nanoscale Imaging',
                'tla':      'ANI',
                'program':  'Imaging and Microscopy',
                'status':   'development'},
    
    '26-ID-1': {'fullname': 'Tender X-ray Nanoprobe',
                'tla':      'TXN',
                'program':  'Imaging and Microscopy',
                'status':   'development'},
    
    '27-ID':   {'fullname': 'High Energy Engineering X-ray Scattering',
                'tla':      'HEX',
                'program':  'Hard X-ray Methods',
                'status':   'operating'},
    
    '28-ID-1': {'fullname': 'Pair Distribution Function',
                'tla':      'PDF',
                'program':  'Hard X-ray Methods',
                'status':   'operating'},
    
    '28-ID-2': {'fullname': 'X-ray Powder Diffraction',
                'tla':      'XPD',
                'program':  'Hard X-ray Methods',
                'status':   'operating'},
    
    '29-ID-1': {'fullname': 'Soft X-ray Nanoprobe',
                'tla':      'SXN',
                'program':  'Imaging and Microscopy',
                'status':   'development'},
    
    '29-ID-2': {'fullname': 'Soft X-ray Photoemission and Scattering Imagine',
                'tla':      'ARI',
                'program':  'Electronic Structure Techniques',
                'status':   'development'},
}

wb = Workbook()
ws = wb.active
ws.append(['port', 'TLA', 'program', 'total', 'high profile', 'years operating',
           '% high profile', 'pubs per year', 'total citations', 'citations per paper'])
c = ws['A2']
ws.freeze_panes = c
for cell in range(20):
    ws[f'{chr(65+cell)}1'].font = Font(bold=True)

i = -1
for j, port in enumerate(beamlines.keys()):
    if beamlines[port]['status'] == 'development':
        continue
    i = i+1
    tla = beamlines[port]['tla']
    program = beamlines[port]['program']
    print(tla)
    url = f'https://www.bnl.gov/nsls2/beamlines/publications.php?q={port}'
    html = requests.get(url).content
    df_list = pd.read_html(html)
    df = df_list[-1]
    answer = str(df).split('\n')[-1].split()
    years = len(str(df).split('\n')) - 2

    matches=re.findall('Cited (\\d+) times', str(html))
    total_citations = sum((int(x) for x in matches))

    ws.append([port, tla, program, int(answer[3]), int(answer[2]), years,
               f'=E{i+3}/D{i+3}',
               f'=D{i+3}/F{i+3}',
               total_citations,
               f'=I{i+3}/D{i+3}',
               ])
    ws[f'G{i+3}'].number_format = '0.00'
    ws[f'H{i+3}'].number_format = '0.00'
    ws[f'J{i+3}'].number_format = '0.00'



ws.column_dimensions['C'].width = 2   *  ws.column_dimensions['D'].width
ws.column_dimensions['E'].width = 1.2 *  ws.column_dimensions['D'].width
ws.column_dimensions['F'].width = 1.3 *  ws.column_dimensions['E'].width
ws.column_dimensions['G'].width = 1.2 *  ws.column_dimensions['F'].width
ws.column_dimensions['H'].width = 1   *  ws.column_dimensions['G'].width
ws.column_dimensions['I'].width = 1   *  ws.column_dimensions['H'].width
ws.column_dimensions['J'].width = 1   *  ws.column_dimensions['H'].width
wb.save("beamline_publication_data.xlsx")
